#This code is written by Mr.Mohammad Amir Jamil. 
#The code bascially is for bulk qr code generator where inputs are fed from Excel file. 
#The output shall be reflected in Excel file as well as images folder.
#Need to work on Code Semantics i.e. variables names since it was made in haste for Project Requirement. 



import qrcode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import time
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

df = pd.read_excel('nation.xlsx', sheet_name='Sheet1',sep='\s*,\s*')
latitudes=df['BUILDINGID'].tolist()
for i in range(0,len(latitudes)):
	print(latitudes[i])
	img = qrcode.make(latitudes[i])
	img.save('images/'+latitudes[i]+'.jpg')
time.sleep(6)
wb = Workbook()
dest_filename = 'final_book.xlsx'
ws2 = wb.create_sheet(title="scancode")
for i in range(1,len(latitudes)+1):
	ws2.row_dimensions[i].height = 250
ws2.column_dimensions['B'].width=45
ws2.column_dimensions['A'].width=45
for j in range(0,len(latitudes)):
	img=Image('images/'+latitudes[j]+'.jpg')
	cellname_image='B'+str(j+1)
	ws2.add_image(img,cellname_image)
	cellname_text='A'+str(j+1)
	ws2[cellname_text]=latitudes[j]
	currentCell = ws2[cellname_text]
	currentCell.alignment = Alignment(horizontal='center',vertical='center')
	currentCell2 = ws2[cellname_image]
	currentCell2.alignment = Alignment(horizontal='center',vertical='center')
	time.sleep(3)


wb.save(filename = dest_filename)
